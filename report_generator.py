import pandas as pd
import numpy as np
import json
import os
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# For Excel generation
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference

# For PDF generation
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# For HTML template rendering
from jinja2 import Template

class ReportGenerator:
    """
    Comprehensive Report Generator for creating Excel, PDF, JSON, and HTML reports
    with detailed technical analysis, price action analysis, and predictions
    """
    
    def __init__(self):
        self.reports_dir = "reports_output"
        self.ensure_reports_directory()
        
    def ensure_reports_directory(self):
        """Ensure reports output directory exists"""
        if not os.path.exists(self.reports_dir):
            os.makedirs(self.reports_dir)
    
    def generate_excel_report(self, market_data, stock_symbol):
        """
        Generate comprehensive multi-sheet Excel report
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.reports_dir}/{stock_symbol}_comprehensive_analysis_{timestamp}.xlsx"
            
            # Create Excel writer object
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                
                # Sheet 1: Executive Summary
                self.create_executive_summary_sheet(writer, market_data, stock_symbol)
                
                # Sheet 2: Raw Market Data
                self.create_market_data_sheet(writer, market_data)
                
                # Sheet 3: Technical Indicators
                self.create_technical_indicators_sheet(writer, market_data)
                
                # Sheet 4: Price Action Analysis
                self.create_price_action_sheet(writer, market_data)
                
                # Sheet 5: Pattern Analysis
                self.create_pattern_analysis_sheet(writer, market_data)
                
                # Sheet 6: Predictions & Forecasts
                self.create_predictions_sheet(writer, market_data, stock_symbol)
                
                # Sheet 7: Risk Analysis
                self.create_risk_analysis_sheet(writer, market_data)
                
                # Sheet 8: Volume Analysis
                self.create_volume_analysis_sheet(writer, market_data)
                
            # Apply formatting to the Excel file
            self.format_excel_file(filename)
            
            return filename
            
        except Exception as e:
            print(f"Excel report generation error: {e}")
            return None
    
    def create_executive_summary_sheet(self, writer, market_data, stock_symbol):
        """Create executive summary sheet"""
        try:
            current_price = market_data['Close'].iloc[-1]
            prev_price = market_data['Close'].iloc[-2]
            price_change = current_price - prev_price
            price_change_pct = (price_change / prev_price) * 100
            
            # Create summary data
            summary_data = {
                'Metric': [
                    'Stock Symbol',
                    'Current Price',
                    'Previous Close',
                    'Price Change',
                    'Price Change %',
                    'Volume',
                    '52-Week High',
                    '52-Week Low',
                    'Market Cap Position',
                    'Volatility (20D)',
                    'Analysis Date',
                    'Recommendation',
                    'Risk Level',
                    'Target Price',
                    'Stop Loss',
                    'Confidence Level'
                ],
                'Value': [
                    stock_symbol,
                    f"₹{current_price:.2f}",
                    f"₹{prev_price:.2f}",
                    f"₹{price_change:.2f}",
                    f"{price_change_pct:.2f}%",
                    f"{market_data['Volume'].iloc[-1]:,.0f}",
                    f"₹{market_data['High'].rolling(252).max().iloc[-1]:.2f}",
                    f"₹{market_data['Low'].rolling(252).min().iloc[-1]:.2f}",
                    self.calculate_market_position(market_data),
                    f"{market_data['Close'].pct_change().rolling(20).std().iloc[-1]*np.sqrt(252)*100:.1f}%",
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    self.generate_overall_recommendation(market_data),
                    self.assess_overall_risk(market_data),
                    f"₹{current_price * 1.05:.2f}",  # Example target
                    f"₹{current_price * 0.95:.2f}",  # Example stop loss
                    "Medium"  # Example confidence
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)
            
        except Exception as e:
            print(f"Executive summary sheet error: {e}")
    
    def create_market_data_sheet(self, writer, market_data):
        """Create raw market data sheet"""
        try:
            # Prepare market data with additional calculated columns
            data_df = market_data.copy()
            
            # Add additional columns
            data_df['Daily_Return'] = data_df['Close'].pct_change()
            data_df['Daily_Range'] = data_df['High'] - data_df['Low']
            data_df['Body_Size'] = abs(data_df['Close'] - data_df['Open'])
            data_df['Upper_Shadow'] = data_df['High'] - data_df[['Open', 'Close']].max(axis=1)
            data_df['Lower_Shadow'] = data_df[['Open', 'Close']].min(axis=1) - data_df['Low']
            
            # Round numerical columns
            numeric_columns = ['Open', 'High', 'Low', 'Close', 'Daily_Return', 'Daily_Range', 'Body_Size', 'Upper_Shadow', 'Lower_Shadow']
            for col in numeric_columns:
                if col in data_df.columns:
                    data_df[col] = data_df[col].round(2)
            
            data_df.to_excel(writer, sheet_name='Market Data', index=True)
            
        except Exception as e:
            print(f"Market data sheet error: {e}")
    
    def create_technical_indicators_sheet(self, writer, market_data):
        """Create technical indicators analysis sheet"""
        try:
            from models.technical_indicators import TechnicalIndicatorEngine
            
            tech_engine = TechnicalIndicatorEngine()
            indicators = tech_engine.calculate_all_indicators(market_data)
            
            # Convert indicators to DataFrame format
            indicators_data = []
            for indicator_name, indicator_data in indicators.items():
                indicators_data.append({
                    'Indicator': indicator_name,
                    'Current_Value': indicator_data.get('current_value', 'N/A'),
                    'Signal': indicator_data.get('signal', 'N/A'),
                    'Pattern': indicator_data.get('pattern_detected', 'N/A'),
                    'Confidence': f"{indicator_data.get('confidence', 0):.1f}%"
                })
            
            indicators_df = pd.DataFrame(indicators_data)
            indicators_df.to_excel(writer, sheet_name='Technical Indicators', index=False)
            
            # Add detailed indicator calculations
            detailed_data = market_data.copy()
            
            # Calculate common technical indicators
            import talib
            try:
                detailed_data['RSI'] = talib.RSI(market_data['Close'].values, timeperiod=14)
                detailed_data['MACD'], detailed_data['MACD_Signal'], detailed_data['MACD_Histogram'] = talib.MACD(market_data['Close'].values)
                detailed_data['BB_Upper'], detailed_data['BB_Middle'], detailed_data['BB_Lower'] = talib.BBANDS(market_data['Close'].values)
                detailed_data['SMA_20'] = talib.SMA(market_data['Close'].values, timeperiod=20)
                detailed_data['EMA_12'] = talib.EMA(market_data['Close'].values, timeperiod=12)
                detailed_data['Stoch_K'], detailed_data['Stoch_D'] = talib.STOCH(market_data['High'].values, market_data['Low'].values, market_data['Close'].values)
            except Exception as talib_error:
                print(f"TA-Lib calculation error: {talib_error}")
            
            # Round numerical columns
            numeric_columns = detailed_data.select_dtypes(include=[np.number]).columns
            detailed_data[numeric_columns] = detailed_data[numeric_columns].round(4)
            
            detailed_data.to_excel(writer, sheet_name='Detailed Indicators', index=True)
            
        except Exception as e:
            print(f"Technical indicators sheet error: {e}")
    
    def create_price_action_sheet(self, writer, market_data):
        """Create price action analysis sheet"""
        try:
            from analysis.price_action_analysis import PriceActionAnalyzer
            
            price_analyzer = PriceActionAnalyzer()
            price_analysis = price_analyzer.analyze_price_action(market_data)
            
            # Market Structure Analysis
            structure_data = price_analysis.get('market_structure', {})
            structure_df = pd.DataFrame([
                ['Structure Type', structure_data.get('structure_type', 'N/A')],
                ['Trend Strength', f"{structure_data.get('trend_strength', 0)}/10"],
                ['Recent Peaks', structure_data.get('recent_peaks', 0)],
                ['Recent Troughs', structure_data.get('recent_troughs', 0)],
                ['Structure Confidence', f"{structure_data.get('structure_confidence', 0):.1f}%"]
            ], columns=['Metric', 'Value'])
            
            structure_df.to_excel(writer, sheet_name='Price Action', index=False, startrow=1)
            
            # Volume-Price Analysis
            volume_data = price_analysis.get('volume_price_analysis', {})
            volume_df = pd.DataFrame([
                ['Volume Trend', volume_data.get('volume_trend', 'N/A')],
                ['Accumulation/Distribution', volume_data.get('accumulation_distribution', 'N/A')],
                ['Volume Confirmation', volume_data.get('volume_confirmation', 'N/A')],
                ['Buying Pressure', f"{volume_data.get('buying_pressure', 50):.1f}%"],
                ['Selling Pressure', f"{volume_data.get('selling_pressure', 50):.1f}%"],
                ['Volume Breakouts', volume_data.get('volume_breakouts', 0)]
            ], columns=['Volume Metric', 'Value'])
            
            volume_df.to_excel(writer, sheet_name='Price Action', index=False, startrow=10)
            
            # Support and Resistance Levels
            sr_data = price_analysis.get('support_resistance', {})
            resistance_levels = sr_data.get('resistance_levels', [])
            support_levels = sr_data.get('support_levels', [])
            
            sr_df = pd.DataFrame({
                'Level Type': ['Resistance'] * len(resistance_levels) + ['Support'] * len(support_levels),
                'Price Level': [f"₹{level:.2f}" for level in resistance_levels + support_levels],
                'Strength': sr_data.get('resistance_strength', [50] * len(resistance_levels)) + sr_data.get('support_strength', [50] * len(support_levels))
            })
            
            sr_df.to_excel(writer, sheet_name='Price Action', index=False, startrow=20)
            
        except Exception as e:
            print(f"Price action sheet error: {e}")
    
    def create_pattern_analysis_sheet(self, writer, market_data):
        """Create pattern analysis sheet"""
        try:
            from models.pattern_discovery import PatternDiscoveryEngine
            
            pattern_engine = PatternDiscoveryEngine()
            patterns = pattern_engine.discover_patterns(market_data)
            
            # Discovered Patterns
            discovered_patterns = patterns.get('discovered_patterns', [])
            if discovered_patterns:
                patterns_df = pd.DataFrame([
                    {
                        'Pattern Name': pattern.get('name', 'Unknown'),
                        'Occurrences': pattern.get('occurrences', 0),
                        'Success Rate': f"{pattern.get('success_rate', 0):.1f}%",
                        'Confidence': f"{pattern.get('confidence', 0):.2f}",
                        'Average Return': f"{pattern.get('avg_return', 0):.3f}"
                    }
                    for pattern in discovered_patterns[:10]  # Top 10 patterns
                ])
                patterns_df.to_excel(writer, sheet_name='Pattern Analysis', index=False)
            
            # Traditional Candlestick Patterns
            traditional_patterns = patterns.get('traditional_candlestick', [])
            if traditional_patterns:
                traditional_df = pd.DataFrame([
                    {
                        'Pattern Name': pattern.get('name', 'Unknown'),
                        'Pattern Type': pattern.get('type', 'Unknown'),
                        'Reliability': f"{pattern.get('reliability', 0):.1f}%",
                        'Timeframe': pattern.get('timeframe', 'Unknown')
                    }
                    for pattern in traditional_patterns[:15]  # Last 15 patterns
                ])
                traditional_df.to_excel(writer, sheet_name='Pattern Analysis', index=False, startrow=len(discovered_patterns) + 3)
            
            # Chart Patterns
            chart_patterns = patterns.get('chart_patterns', [])
            if chart_patterns:
                chart_df = pd.DataFrame([
                    {
                        'Chart Pattern': pattern.get('pattern_type', 'Unknown'),
                        'Formation Stage': pattern.get('formation_stage', 'Unknown'),
                        'Completion Probability': f"{pattern.get('completion_probability', 0):.2f}",
                        'Expected Breakout': pattern.get('expected_breakout', 'Unknown'),
                        'Price Target': f"₹{pattern.get('price_target', 0):.2f}"
                    }
                    for pattern in chart_patterns[:10]
                ])
                chart_df.to_excel(writer, sheet_name='Pattern Analysis', index=False, startrow=len(discovered_patterns) + len(traditional_patterns) + 6)
            
        except Exception as e:
            print(f"Pattern analysis sheet error: {e}")
    
    def create_predictions_sheet(self, writer, market_data, stock_symbol):
        """Create predictions and forecasts sheet"""
        try:
            from models.prediction_engine import PredictionEngine
            
            pred_engine = PredictionEngine()
            predictions = pred_engine.generate_predictions(market_data)
            
            # Primary Prediction
            primary_pred = predictions.get('primary_prediction', {})
            primary_df = pd.DataFrame([
                ['Direction', primary_pred.get('direction', 'Neutral')],
                ['Probability', f"{primary_pred.get('probability', 50):.1f}%"],
                ['Target Range Low', f"₹{primary_pred.get('target_range', {}).get('low', 0):.2f}"],
                ['Target Range High', f"₹{primary_pred.get('target_range', {}).get('high', 0):.2f}"],
                ['Time Horizon', primary_pred.get('time_horizon', 'Unknown')],
                ['Confidence Level', primary_pred.get('confidence_level', 'Low')],
                ['Confidence Score', f"{primary_pred.get('confidence', 0):.1f}%"]
            ], columns=['Prediction Metric', 'Value'])
            
            primary_df.to_excel(writer, sheet_name='Predictions', index=False)
            
            # Scenario Analysis
            scenarios = predictions.get('scenario_analysis', {})
            scenario_data = []
            for scenario_name, scenario_info in scenarios.items():
                scenario_data.append({
                    'Scenario': scenario_name,
                    'Price': f"₹{scenario_info.get('price', 0):.2f}",
                    'Return': f"{scenario_info.get('return', 0):.1f}%",
                    'Probability': f"{scenario_info.get('probability', 0):.1f}%"
                })
            
            if scenario_data:
                scenario_df = pd.DataFrame(scenario_data)
                scenario_df.to_excel(writer, sheet_name='Predictions', index=False, startrow=10)
            
            # Feature Importance
            feature_importance = predictions.get('feature_importance', {})
            if feature_importance:
                feature_df = pd.DataFrame([
                    {'Feature': feature, 'Importance': f"{importance:.3f}"}
                    for feature, importance in feature_importance.items()
                ])
                feature_df.to_excel(writer, sheet_name='Predictions', index=False, startrow=17)
            
        except Exception as e:
            print(f"Predictions sheet error: {e}")
    
    def create_risk_analysis_sheet(self, writer, market_data):
        """Create risk analysis sheet"""
        try:
            # Calculate risk metrics
            returns = market_data['Close'].pct_change().dropna()
            
            risk_metrics = {
                'Metric': [
                    'Daily Volatility',
                    'Annualized Volatility',
                    'VaR (95%)',
                    'Maximum Drawdown',
                    'Sharpe Ratio (Approx)',
                    'Beta (Market)',
                    'Current Risk Level',
                    'Position Size Recommendation',
                    'Stop Loss Distance',
                    'Risk-Reward Ratio'
                ],
                'Value': [
                    f"{returns.std():.4f}",
                    f"{returns.std() * np.sqrt(252):.2f}",
                    f"{np.percentile(returns, 5):.4f}",
                    f"{self.calculate_max_drawdown(market_data):.2f}%",
                    f"{self.calculate_sharpe_ratio(returns):.2f}",
                    "N/A",  # Would need market index data
                    self.assess_current_risk_level(market_data),
                    self.recommend_position_size(market_data),
                    f"{self.calculate_stop_loss_distance(market_data):.2f}%",
                    "1:2"  # Example ratio
                ]
            }
            
            risk_df = pd.DataFrame(risk_metrics)
            risk_df.to_excel(writer, sheet_name='Risk Analysis', index=False)
            
            # Historical volatility analysis
            volatility_periods = [5, 10, 20, 50, 100]
            vol_data = []
            for period in volatility_periods:
                if len(returns) >= period:
                    vol = returns.tail(period).std() * np.sqrt(252)
                    vol_data.append({
                        'Period': f'{period} Days',
                        'Volatility': f'{vol:.2f}',
                        'Risk Level': 'High' if vol > 0.4 else 'Medium' if vol > 0.25 else 'Low'
                    })
            
            if vol_data:
                vol_df = pd.DataFrame(vol_data)
                vol_df.to_excel(writer, sheet_name='Risk Analysis', index=False, startrow=15)
            
        except Exception as e:
            print(f"Risk analysis sheet error: {e}")
    
    def create_volume_analysis_sheet(self, writer, market_data):
        """Create volume analysis sheet"""
        try:
            # Volume statistics
            volume_stats = {
                'Metric': [
                    'Current Volume',
                    'Average Volume (20D)',
                    'Average Volume (50D)',
                    'Volume Ratio (Current/20D)',
                    'Highest Volume (50D)',
                    'Lowest Volume (50D)',
                    'Volume Trend',
                    'Volume Volatility'
                ],
                'Value': [
                    f"{market_data['Volume'].iloc[-1]:,.0f}",
                    f"{market_data['Volume'].rolling(20).mean().iloc[-1]:,.0f}",
                    f"{market_data['Volume'].rolling(50).mean().iloc[-1]:,.0f}",
                    f"{market_data['Volume'].iloc[-1] / market_data['Volume'].rolling(20).mean().iloc[-1]:.2f}",
                    f"{market_data['Volume'].rolling(50).max().iloc[-1]:,.0f}",
                    f"{market_data['Volume'].rolling(50).min().iloc[-1]:,.0f}",
                    self.analyze_volume_trend(market_data),
                    f"{market_data['Volume'].pct_change().rolling(20).std().iloc[-1]:.4f}"
                ]
            }
            
            volume_df = pd.DataFrame(volume_stats)
            volume_df.to_excel(writer, sheet_name='Volume Analysis', index=False)
            
            # Volume-Price correlation analysis
            price_changes = market_data['Close'].pct_change()
            volume_changes = market_data['Volume'].pct_change()
            
            correlation_data = []
            periods = [5, 10, 20, 50]
            for period in periods:
                if len(price_changes) >= period:
                    corr = price_changes.tail(period).corr(volume_changes.tail(period))
                    correlation_data.append({
                        'Period': f'{period} Days',
                        'Price-Volume Correlation': f'{corr:.3f}',
                        'Interpretation': 'Positive' if corr > 0.1 else 'Negative' if corr < -0.1 else 'Neutral'
                    })
            
            if correlation_data:
                corr_df = pd.DataFrame(correlation_data)
                corr_df.to_excel(writer, sheet_name='Volume Analysis', index=False, startrow=12)
            
        except Exception as e:
            print(f"Volume analysis sheet error: {e}")
    
    def format_excel_file(self, filename):
        """Apply professional formatting to Excel file"""
        try:
            workbook = openpyxl.load_workbook(filename)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Format headers
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(filename)
            
        except Exception as e:
            print(f"Excel formatting error: {e}")
    
    def generate_pdf_reports(self, market_data, stock_symbol):
        """
        Generate separate PDF reports for Technical and Price Action analysis
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Generate Technical Analysis PDF
            tech_filename = f"{self.reports_dir}/{stock_symbol}_technical_analysis_{timestamp}.pdf"
            self.create_technical_analysis_pdf(market_data, stock_symbol, tech_filename)
            
            # Generate Price Action Analysis PDF
            price_filename = f"{self.reports_dir}/{stock_symbol}_price_action_analysis_{timestamp}.pdf"
            self.create_price_action_pdf(market_data, stock_symbol, price_filename)
            
            return {
                'Technical Analysis': tech_filename,
                'Price Action Analysis': price_filename
            }
            
        except Exception as e:
            print(f"PDF reports generation error: {e}")
            return {}
    
    def create_technical_analysis_pdf(self, market_data, stock_symbol, filename):
        """Create Technical Analysis PDF report"""
        try:
            doc = SimpleDocTemplate(filename, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], alignment=TA_CENTER, spaceAfter=30)
            story.append(Paragraph(f"Technical Analysis Report - {stock_symbol}", title_style))
            story.append(Spacer(1, 20))
            
            # Current market status
            current_price = market_data['Close'].iloc[-1]
            prev_price = market_data['Close'].iloc[-2]
            change = current_price - prev_price
            change_pct = (change / prev_price) * 100
            
            story.append(Paragraph(f"<b>Current Price:</b> ₹{current_price:.2f} ({change:+.2f}, {change_pct:+.2f}%)", styles['Normal']))
            story.append(Paragraph(f"<b>Analysis Date:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Technical Indicators Summary
            story.append(Paragraph("Technical Indicators Analysis", styles['Heading2']))
            
            try:
                from models.technical_indicators import TechnicalIndicatorEngine
                tech_engine = TechnicalIndicatorEngine()
                indicators = tech_engine.calculate_all_indicators(market_data)
                
                # Create table data for key indicators
                table_data = [['Indicator', 'Value', 'Signal', 'Confidence']]
                
                key_indicators = ['RSI', 'MACD', 'Bollinger_Bands', 'Stochastic', 'Volume_OBV']
                for indicator in key_indicators:
                    if indicator in indicators:
                        data = indicators[indicator]
                        table_data.append([
                            indicator.replace('_', ' '),
                            str(data.get('current_value', 'N/A'))[:10],
                            data.get('signal', 'N/A'),
                            f"{data.get('confidence', 0):.1f}%"
                        ])
                
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(table)
                
            except Exception as e:
                story.append(Paragraph(f"Technical indicators analysis unavailable: {e}", styles['Normal']))
            
            story.append(Spacer(1, 20))
            
            # Support and Resistance Analysis
            story.append(Paragraph("Support & Resistance Analysis", styles['Heading2']))
            
            try:
                from analysis.technical_analysis import TechnicalAnalyzer
                tech_analyzer = TechnicalAnalyzer()
                tech_analysis = tech_analyzer.generate_comprehensive_analysis(market_data)
                
                sr_data = tech_analysis.get('support_resistance', {})
                resistance_levels = sr_data.get('resistance', [])
                support_levels = sr_data.get('support', [])
                
                story.append(Paragraph(f"<b>Key Resistance Levels:</b> {', '.join([f'₹{r:.2f}' for r in resistance_levels[:3]])}", styles['Normal']))
                story.append(Paragraph(f"<b>Key Support Levels:</b> {', '.join([f'₹{s:.2f}' for s in support_levels[:3]])}", styles['Normal']))
                
            except Exception as e:
                story.append(Paragraph(f"Support/Resistance analysis unavailable: {e}", styles['Normal']))
            
            story.append(Spacer(1, 20))
            
            # Risk Assessment
            story.append(Paragraph("Risk Assessment", styles['Heading2']))
            
            returns = market_data['Close'].pct_change().dropna()
            volatility = returns.std() * np.sqrt(252)
            
            risk_level = "High" if volatility > 0.4 else "Medium" if volatility > 0.25 else "Low"
            story.append(Paragraph(f"<b>Volatility:</b> {volatility:.2f} ({risk_level} Risk)", styles['Normal']))
            story.append(Paragraph(f"<b>Recommended Position Size:</b> {self.recommend_position_size(market_data)}", styles['Normal']))
            
            # Build PDF
            doc.build(story)
            
        except Exception as e:
            print(f"Technical PDF creation error: {e}")
    
    def create_price_action_pdf(self, market_data, stock_symbol, filename):
        """Create Price Action Analysis PDF report"""
        try:
            doc = SimpleDocTemplate(filename, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], alignment=TA_CENTER, spaceAfter=30)
            story.append(Paragraph(f"Price Action Analysis Report - {stock_symbol}", title_style))
            story.append(Spacer(1, 20))
            
            # Market Structure Analysis
            story.append(Paragraph("Market Structure Analysis", styles['Heading2']))
            
            try:
                from analysis.price_action_analysis import PriceActionAnalyzer
                price_analyzer = PriceActionAnalyzer()
                price_analysis = price_analyzer.analyze_price_action(market_data)
                
                structure = price_analysis.get('market_structure', {})
                story.append(Paragraph(f"<b>Structure Type:</b> {structure.get('structure_type', 'Unknown')}", styles['Normal']))
                story.append(Paragraph(f"<b>Trend Strength:</b> {structure.get('trend_strength', 0)}/10", styles['Normal']))
                story.append(Paragraph(f"<b>Confidence:</b> {structure.get('structure_confidence', 0):.1f}%", styles['Normal']))
                
            except Exception as e:
                story.append(Paragraph(f"Market structure analysis unavailable: {e}", styles['Normal']))
            
            story.append(Spacer(1, 20))
            
            # Volume-Price Relationship
            story.append(Paragraph("Volume-Price Relationship", styles['Heading2']))
            
            try:
                volume_analysis = price_analysis.get('volume_price_analysis', {})
                story.append(Paragraph(f"<b>Volume Trend:</b> {volume_analysis.get('volume_trend', 'Unknown')}", styles['Normal']))
                story.append(Paragraph(f"<b>Accumulation/Distribution:</b> {volume_analysis.get('accumulation_distribution', 'Unknown')}", styles['Normal']))
                story.append(Paragraph(f"<b>Buying Pressure:</b> {volume_analysis.get('buying_pressure', 50):.1f}%", styles['Normal']))
                story.append(Paragraph(f"<b>Selling Pressure:</b> {volume_analysis.get('selling_pressure', 50):.1f}%", styles['Normal']))
                
            except:
                story.append(Paragraph("Volume-price analysis unavailable", styles['Normal']))
            
            story.append(Spacer(1, 20))
            
            # Pattern Analysis
            story.append(Paragraph("Candlestick Pattern Analysis", styles['Heading2']))
            
            try:
                from models.pattern_discovery import PatternDiscoveryEngine
                pattern_engine = PatternDiscoveryEngine()
                patterns = pattern_engine.detect_candlestick_patterns(market_data)
                
                traditional_patterns = patterns.get('traditional_patterns', [])
                if traditional_patterns:
                    story.append(Paragraph("<b>Recent Patterns Detected:</b>", styles['Normal']))
                    for pattern in traditional_patterns[-5:]:  # Last 5 patterns
                        story.append(Paragraph(f"• {pattern.get('name', 'Unknown')} ({pattern.get('type', 'Unknown')}) - Reliability: {pattern.get('reliability', 0):.1f}%", styles['Normal']))
                else:
                    story.append(Paragraph("No significant patterns detected recently", styles['Normal']))
                
            except Exception as e:
                story.append(Paragraph(f"Pattern analysis unavailable: {e}", styles['Normal']))
            
            story.append(Spacer(1, 20))
            
            # Entry/Exit Points
            story.append(Paragraph("Entry & Exit Analysis", styles['Heading2']))
            
            try:
                entry_exit = price_analysis.get('entry_exit_points', {})
                entry_points = entry_exit.get('entry_points', {})
                exit_points = entry_exit.get('exit_points', {})
                
                story.append(Paragraph(f"<b>Entry Bias:</b> {entry_points.get('entry_bias', 'Neutral')}", styles['Normal']))
                story.append(Paragraph(f"<b>Long Entry:</b> ₹{entry_points.get('long_entry', 0):.2f}", styles['Normal']))
                story.append(Paragraph(f"<b>Short Entry:</b> ₹{entry_points.get('short_entry', 0):.2f}", styles['Normal']))
                story.append(Paragraph(f"<b>Stop Loss Long:</b> ₹{exit_points.get('stop_loss_long', 0):.2f}", styles['Normal']))
                story.append(Paragraph(f"<b>Stop Loss Short:</b> ₹{exit_points.get('stop_loss_short', 0):.2f}", styles['Normal']))
                
            except:
                story.append(Paragraph("Entry/Exit analysis unavailable", styles['Normal']))
            
            # Build PDF
            doc.build(story)
            
        except Exception as e:
            print(f"Price Action PDF creation error: {e}")
    
    def generate_json_export(self, market_data, stock_symbol):
        """
        Generate comprehensive JSON export of all analysis data
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.reports_dir}/{stock_symbol}_data_export_{timestamp}.json"
            
            # Collect all analysis data
            export_data = {
                'metadata': {
                    'stock_symbol': stock_symbol,
                    'analysis_timestamp': datetime.now().isoformat(),
                    'data_period': {
                        'start_date': market_data.index[0].isoformat() if not market_data.empty else None,
                        'end_date': market_data.index[-1].isoformat() if not market_data.empty else None,
                        'total_periods': len(market_data)
                    }
                },
                'market_data': {
                    'current_price': float(market_data['Close'].iloc[-1]) if not market_data.empty else 0,
                    'previous_close': float(market_data['Close'].iloc[-2]) if len(market_data) > 1 else 0,
                    'volume': int(market_data['Volume'].iloc[-1]) if not market_data.empty else 0,
                    'high_52w': float(market_data['High'].rolling(252).max().iloc[-1]) if len(market_data) >= 252 else 0,
                    'low_52w': float(market_data['Low'].rolling(252).min().iloc[-1]) if len(market_data) >= 252 else 0
                },
                'technical_analysis': {},
                'price_action_analysis': {},
                'pattern_analysis': {},
                'predictions': {},
                'risk_analysis': {}
            }
            
            # Add technical analysis
            try:
                from models.technical_indicators import TechnicalIndicatorEngine
                from analysis.technical_analysis import TechnicalAnalyzer
                
                tech_engine = TechnicalIndicatorEngine()
                tech_analyzer = TechnicalAnalyzer()
                
                indicators = tech_engine.calculate_all_indicators(market_data)
                tech_analysis = tech_analyzer.generate_comprehensive_analysis(market_data)
                
                export_data['technical_analysis'] = {
                    'indicators': indicators,
                    'comprehensive_analysis': tech_analysis
                }
                
            except Exception as e:
                export_data['technical_analysis'] = {'error': str(e)}
            
            # Add price action analysis
            try:
                from analysis.price_action_analysis import PriceActionAnalyzer
                
                price_analyzer = PriceActionAnalyzer()
                price_analysis = price_analyzer.analyze_price_action(market_data)
                
                export_data['price_action_analysis'] = price_analysis
                
            except Exception as e:
                export_data['price_action_analysis'] = {'error': str(e)}
            
            # Add pattern analysis
            try:
                from models.pattern_discovery import PatternDiscoveryEngine
                
                pattern_engine = PatternDiscoveryEngine()
                patterns = pattern_engine.discover_patterns(market_data)
                
                export_data['pattern_analysis'] = patterns
                
            except Exception as e:
                export_data['pattern_analysis'] = {'error': str(e)}
            
            # Add predictions
            try:
                from models.prediction_engine import PredictionEngine
                
                pred_engine = PredictionEngine()
                predictions = pred_engine.generate_predictions(market_data)
                
                export_data['predictions'] = predictions
                
            except Exception as e:
                export_data['predictions'] = {'error': str(e)}
            
            # Add risk analysis
            try:
                returns = market_data['Close'].pct_change().dropna()
                export_data['risk_analysis'] = {
                    'volatility_daily': float(returns.std()),
                    'volatility_annual': float(returns.std() * np.sqrt(252)),
                    'max_drawdown': self.calculate_max_drawdown(market_data),
                    'sharpe_ratio': self.calculate_sharpe_ratio(returns),
                    'var_95': float(np.percentile(returns, 5)),
                    'risk_level': self.assess_current_risk_level(market_data)
                }
                
            except Exception as e:
                export_data['risk_analysis'] = {'error': str(e)}
            
            # Convert numpy types to Python native types for JSON serialization
            export_data = self.convert_numpy_types(export_data)
            
            # Save to JSON file
            with open(filename, 'w') as f:
                json.dump(export_data, f, indent=2, default=str)
            
            return filename
            
        except Exception as e:
            print(f"JSON export error: {e}")
            return None
    
    def generate_executive_summary(self, market_data, stock_symbol):
        """
        Generate executive summary with key insights and recommendations
        """
        try:
            current_price = market_data['Close'].iloc[-1]
            prev_price = market_data['Close'].iloc[-2]
            price_change_pct = ((current_price - prev_price) / prev_price) * 100
            
            # Overall rating calculation
            returns = market_data['Close'].pct_change().dropna()
            volatility = returns.std() * np.sqrt(252)
            
            # Simple scoring system
            price_momentum_score = 1 if price_change_pct > 2 else 0 if price_change_pct > -2 else -1
            volatility_score = -1 if volatility > 0.4 else 0 if volatility > 0.25 else 1
            volume_score = 1 if market_data['Volume'].iloc[-1] > market_data['Volume'].rolling(20).mean().iloc[-1] else 0
            
            total_score = price_momentum_score + volatility_score + volume_score
            
            if total_score >= 2:
                overall_rating = "Strong Buy"
            elif total_score >= 1:
                overall_rating = "Buy"
            elif total_score >= 0:
                overall_rating = "Hold"
            elif total_score >= -1:
                overall_rating = "Sell"
            else:
                overall_rating = "Strong Sell"
            
            # Generate key findings
            key_findings = []
            
            if abs(price_change_pct) > 3:
                key_findings.append(f"Significant price movement: {price_change_pct:+.1f}% in latest session")
            
            if volatility > 0.4:
                key_findings.append("High volatility environment - increased risk")
            elif volatility < 0.2:
                key_findings.append("Low volatility environment - stable conditions")
            
            volume_ratio = market_data['Volume'].iloc[-1] / market_data['Volume'].rolling(20).mean().iloc[-1]
            if volume_ratio > 1.5:
                key_findings.append("Above-average volume activity detected")
            elif volume_ratio < 0.5:
                key_findings.append("Below-average volume activity")
            
            # 52-week position
            high_52w = market_data['High'].rolling(252).max().iloc[-1]
            low_52w = market_data['Low'].rolling(252).min().iloc[-1]
            position_52w = (current_price - low_52w) / (high_52w - low_52w) * 100
            
            if position_52w > 80:
                key_findings.append("Trading near 52-week highs")
            elif position_52w < 20:
                key_findings.append("Trading near 52-week lows")
            
            # Action items
            action_items = []
            
            if overall_rating in ["Strong Buy", "Buy"]:
                action_items.append("Consider initiating or adding to position")
                action_items.append(f"Set stop loss at ₹{current_price * 0.95:.2f}")
            elif overall_rating in ["Strong Sell", "Sell"]:
                action_items.append("Consider reducing or closing position")
                action_items.append("Monitor for oversold bounce opportunities")
            else:
                action_items.append("Maintain current position")
                action_items.append("Watch for breakout signals")
            
            action_items.append("Monitor volume for confirmation")
            action_items.append("Review position sizing based on volatility")
            
            return {
                'overall_rating': overall_rating,
                'primary_recommendation': f"{overall_rating} - Target: ₹{current_price * 1.05:.2f}",
                'risk_level': "High" if volatility > 0.4 else "Medium" if volatility > 0.25 else "Low",
                'time_horizon': "3-7 trading days",
                'key_findings': key_findings,
                'action_items': action_items,
                'confidence_score': f"{min(80, max(40, 60 + abs(total_score) * 10)):.0f}%"
            }
            
        except Exception as e:
            print(f"Executive summary error: {e}")
            return {
                'overall_rating': 'Hold',
                'primary_recommendation': 'Hold - Insufficient data for strong recommendation',
                'risk_level': 'Medium',
                'time_horizon': '3-7 trading days',
                'key_findings': ['Analysis data incomplete'],
                'action_items': ['Monitor market conditions'],
                'confidence_score': '50%'
            }
    
    # Helper methods
    def calculate_market_position(self, data):
        """Calculate current position relative to 52-week range"""
        try:
            current_price = data['Close'].iloc[-1]
            high_52w = data['High'].rolling(252).max().iloc[-1]
            low_52w = data['Low'].rolling(252).min().iloc[-1]
            position = (current_price - low_52w) / (high_52w - low_52w) * 100
            
            if position > 80:
                return "Near 52W High"
            elif position < 20:
                return "Near 52W Low"
            else:
                return f"{position:.0f}% of 52W Range"
        except:
            return "Unknown"
    
    def generate_overall_recommendation(self, data):
        """Generate overall recommendation"""
        try:
            returns = data['Close'].pct_change().dropna()
            recent_return = returns.tail(5).mean()
            
            if recent_return > 0.02:
                return "Strong Buy"
            elif recent_return > 0.005:
                return "Buy"
            elif recent_return > -0.005:
                return "Hold"
            elif recent_return > -0.02:
                return "Sell"
            else:
                return "Strong Sell"
        except:
            return "Hold"
    
    def assess_overall_risk(self, data):
        """Assess overall risk level"""
        try:
            returns = data['Close'].pct_change().dropna()
            volatility = returns.std() * np.sqrt(252)
            
            if volatility > 0.4:
                return "High Risk"
            elif volatility > 0.25:
                return "Medium Risk"
            else:
                return "Low Risk"
        except:
            return "Medium Risk"
    
    def calculate_max_drawdown(self, data):
        """Calculate maximum drawdown"""
        try:
            prices = data['Close']
            peak = prices.expanding().max()
            drawdown = (prices - peak) / peak * 100
            return abs(drawdown.min())
        except:
            return 0
    
    def calculate_sharpe_ratio(self, returns, risk_free_rate=0.05):
        """Calculate Sharpe ratio"""
        try:
            excess_returns = returns.mean() * 252 - risk_free_rate
            volatility = returns.std() * np.sqrt(252)
            return excess_returns / volatility if volatility > 0 else 0
        except:
            return 0
    
    def assess_current_risk_level(self, data):
        """Assess current risk level"""
        try:
            returns = data['Close'].pct_change().dropna()
            recent_vol = returns.tail(20).std() * np.sqrt(252)
            
            if recent_vol > 0.4:
                return "High"
            elif recent_vol > 0.25:
                return "Medium"
            else:
                return "Low"
        except:
            return "Medium"
    
    def recommend_position_size(self, data):
        """Recommend position size based on volatility"""
        try:
            returns = data['Close'].pct_change().dropna()
            volatility = returns.std() * np.sqrt(252)
            
            if volatility > 0.4:
                return "Small (1-2% of portfolio)"
            elif volatility > 0.25:
                return "Medium (3-5% of portfolio)"
            else:
                return "Large (5-8% of portfolio)"
        except:
            return "Medium (3-5% of portfolio)"
    
    def calculate_stop_loss_distance(self, data):
        """Calculate recommended stop loss distance"""
        try:
            atr = (data['High'] - data['Low']).tail(14).mean()
            current_price = data['Close'].iloc[-1]
            return (atr * 2 / current_price) * 100
        except:
            return 5.0
    
    def analyze_volume_trend(self, data):
        """Analyze volume trend"""
        try:
            volume_ma_short = data['Volume'].rolling(10).mean().iloc[-1]
            volume_ma_long = data['Volume'].rolling(30).mean().iloc[-1]
            
            if volume_ma_short > volume_ma_long * 1.2:
                return "Rising"
            elif volume_ma_short < volume_ma_long * 0.8:
                return "Falling"
            else:
                return "Stable"
        except:
            return "Unknown"
    
    def convert_numpy_types(self, obj):
        """Convert numpy types to Python native types for JSON serialization"""
        if isinstance(obj, dict):
            return {k: self.convert_numpy_types(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [self.convert_numpy_types(item) for item in obj]
        elif isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif pd.isna(obj):
            return None
        else:
            return obj
